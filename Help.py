import os.path

from openpyxl import load_workbook


def parsing_timesheet(filename):
    # грузим книгу
    wb = load_workbook(os.path.join(filename), read_only=True)
    # выбираем лист
    ws = wb['Лист1']
    translate = {"Кл час": 'Классный час', 'АЛГ': 'Алгебра', 'АНГяз': 'Английский язык', 'ЛИТ': 'Литература',
                 'ФИЗ': 'Физика', 'ИНФ': 'Информатика', 'РУС': 'Русский язык', 'МУЗ': 'Музыка', 'ИСТ': 'История',
                 'ТЕХ/1': 'Технология (1 группа)', 'Ф-ра': 'Физичекская культура', 'БАРН': 'Барнауловедение', 'ИЗО': 'ИЗО',
                 'БИО': 'Биология', 'ГЕОГ': 'География', 'ГЕОМ': 'Геометрия', 'ОБЩ': 'Обществознание',
                 'ВиС': 'Вероятность и статистика', 'ТЕХ/2': 'Технология (2 группа)', 'ХИМ': 'Химия', 'ОБЖ': 'ОБЖ',
                 'МАТэ': 'Электив по математике', 'РУСэ': 'Электив по русскому языку', 'ФинГ': 'Финансовая грамотность',
                 'ИСТэ': 'Электив по истории', 'ТЕХ': 'Технология', 'АНГяз-1': 'Английский язык (1 группа)',
                 'ИНФэ': 'Электив по информатике', 'ФИЗ-1ч': 'Физика', 'ФИЗ-2ч': 'Физика',
                 'ОБЩэ': 'Электив по обществознанию', 'УПП': 'Проектная деятельность', 'Псих.': 'Психология',
                 'МАТ': 'Математика', 'АнгЯз-2': 'Английский язык (2 группа)', 'Пусто': 'Пусто'}

    days_of_week = ['Понедельник', "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]
    timesheet = {}
    day = {}
    lesson = {}
    # осальные сроки
    count = 0
    for column in range(4, ws.max_column, 2):
        for row in range(7, 70):
            if row in range(34, 37):
                continue
            try:
                count += 1
                if count == 9:
                    day[days_of_week[len(day.keys())]] = lesson
                    count = 0
                    lesson = {}
                else:
                    lesson[str(count)] = []
                    if ws.cell(row, column).value is None and len(lesson.keys()) > 0:
                        lesson[str(count)].append('Пусто')
                    else:
                        lesson[str(count)].append(translate[ws.cell(row, column).value])
                        lesson[str(count)].append(ws.cell(row, column + 1).value)
            except Exception:
                continue
        timesheet[ws.cell(6, column).value] = day
        day = {}
        count = 0
    return timesheet


