from openpyxl import load_workbook

# грузим книгу
wb = load_workbook('test.xlsx', read_only=True)
# выбираем лист
ws = wb['Лист1']
# печааем заголовок
print(f'{ws.cell(1,1).value}\t{ws.cell(1,2).value}\t{ws.cell(1,3).value}')
# осальные сроки
for row in range(2,ws.max_row):
    print(f'{ws.cell(row,1).value}\t{ws.cell(row,2).value}\t{ws.cell(row,3).value}')
